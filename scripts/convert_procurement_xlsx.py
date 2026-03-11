"""
Convert Dataset-Procurement Analysis Sample.xlsx to a flat CSV
that the vendor-mcp pipeline can ingest.

Join: invoice line item fact → invoice (for date) → vendor → item (for category)
Use: Invoice Amount (USD-converted) where available, else Unit Price LOC * Quantity
"""
import sys
import csv
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


def parse_amount(val) -> float:
    """Parse European or US formatted dollar strings like $18.494,28 or $1,234.56"""
    if val is None:
        return 0.0
    s = str(val).strip().lstrip("$").replace(" ", "")
    # Detect European format: period as thousands sep, comma as decimal
    if re.search(r'\.\d{3},', s) or (s.count('.') > 1) or (s.count(',') == 1 and s.count('.') > 0 and s.index(',') > s.rindex('.')):
        s = s.replace('.', '').replace(',', '.')
    else:
        # US format: comma as thousands sep
        s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def main():
    src = Path(__file__).parent.parent / "Dataset-Procurement Analysis Sample.xlsx"
    if not src.exists():
        print(f"File not found: {src}")
        sys.exit(1)

    out = Path(__file__).parent.parent / "data" / "incoming" / "procurement_sample.csv"
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading {src.name} ...")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    # ── Load vendor lookup: VendorID → Vendor Name ──────────────────────────
    print("  Loading vendors ...")
    vendors = {}
    for row in wb["vendor"].iter_rows(min_row=2, values_only=True):
        vid, city, state, postal, country, total_spend, tier, name = row
        if vid and name:
            vendors[vid] = str(name).strip()

    # ── Load item lookup: ItemID → Category, Sub Category ───────────────────
    print("  Loading items ...")
    items = {}
    for row in wb["item "].iter_rows(min_row=2, values_only=True):
        iid, category, sub_cat, commodity, item, detail = row
        if iid:
            items[iid] = {
                "category": str(category or "Other").strip(),
                "sub_category": str(sub_cat or "").strip(),
            }

    # ── Load invoice lookup: InvoiceID → VendorID, LocationID ───────────────
    print("  Loading invoices ...")
    invoices = {}
    for row in wb["invoice"].iter_rows(min_row=2, values_only=True):
        inv_id, date_id, vendor_id, loc_id, disc_days, disc_pct, pay_terms = row
        if inv_id:
            invoices[inv_id] = {"vendor_id": vendor_id, "date_id": date_id}

    # ── Load date lookup: DateID → Year, Month, Day ─────────────────────────
    print("  Loading dates ...")
    dates = {}
    for row in wb["date"].iter_rows(min_row=2, values_only=True):
        # DateID, Day, DayOfWeekNo, DayOfWeek, DayOfYear, WeekNo, MonthNo, Month, QtrNo, Quarter, Year, Mon
        date_id = row[0]
        day = row[1]
        month_no = row[6]
        year = row[10]
        if date_id and year:
            dates[date_id] = f"{year}-{str(month_no).zfill(2)}-{str(day).zfill(2)}"

    # ── Stream line items and write CSV ─────────────────────────────────────
    print("  Processing line items ...")
    written = 0
    skipped = 0

    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["TransactionID", "Vendor", "Category", "SubCategory",
                         "Amount_USD", "Date", "InvoiceID"])

        for i, row in enumerate(wb["invoice line item fact"].iter_rows(min_row=2, values_only=True)):
            date_id, inv_id, item_id, currency_id, line_item, qty, unit_price_loc, inv_loc_amt, date_currency_id, exchange_rate, inv_amount_usd, savings = row

            # Get USD amount: prefer Invoice Amount (already converted), else skip
            amount = parse_amount(inv_amount_usd)
            if amount <= 0:
                # Fallback: use LOC amount — only reliable if USD
                amount = parse_amount(inv_loc_amt)
            if amount <= 0:
                skipped += 1
                continue

            inv = invoices.get(inv_id, {})
            vendor_id = inv.get("vendor_id")
            vendor_name = vendors.get(vendor_id, f"Unknown-{vendor_id}")

            item = items.get(item_id, {"category": "Other", "sub_category": ""})
            date_str = dates.get(date_id, "")

            writer.writerow([
                f"LI{i}",
                vendor_name,
                item["category"],
                item["sub_category"],
                f"{amount:.2f}",
                date_str,
                inv_id,
            ])
            written += 1

            if written % 50000 == 0:
                print(f"    {written:,} rows written ...")

    print(f"\nDone. {written:,} rows written, {skipped:,} skipped (no USD amount).")
    print(f"Output: {out}")


if __name__ == "__main__":
    main()
